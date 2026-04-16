"""
╔══════════════════════════════════════════════════════════╗
║         ZALO GROUP SCANNER - GUI v2.0                    ║
║  Multi-account · Excel Export · Friend Request           ║
║  Group Invite · Admin Filter                             ║
╚══════════════════════════════════════════════════════════╝
"""

import sys, os, json, subprocess, threading, csv, time, re, random, shutil
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XLImage
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import requests
    from io import BytesIO
    HAS_REQUESTS = True
except ImportError:
    HAS_REQUESTS = False

try:
    from PIL import Image as PILImage
    HAS_PILLOW = True
except ImportError:
    HAS_PILLOW = False

BASE_DIR = Path(__file__).parent
CREDENTIALS_FILE = BASE_DIR / "credentials.json"
EXCEL_DATA_DIR = BASE_DIR / "excel_data"
EXCEL_DATA_DIR.mkdir(exist_ok=True)
AVATAR_CACHE_DIR = BASE_DIR / "avatar_cache"
AVATAR_CACHE_DIR.mkdir(exist_ok=True)
LAST_SESSION_FILE = BASE_DIR / "last_session.json"
TEMPLATES_FILE = BASE_DIR / "templates.json"
IMAGE_DIR = BASE_DIR / "anh_ket_ban_nhan_tin"
IMAGE_DIR.mkdir(exist_ok=True)

DEFAULT_USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/130.0.0.0 Safari/537.36"
)

# ============================================================
# THEME
# ============================================================
C = {
    "bg": "#0d1117", "card": "#161b22", "card2": "#1c2333",
    "input": "#21262d", "hover": "#30363d", "border": "#30363d",
    "accent": "#58a6ff", "green": "#3fb950", "orange": "#f78166",
    "purple": "#d2a8ff", "red": "#f85149", "yellow": "#d29922",
    "text": "#e6edf3", "dim": "#8b949e", "bright": "#ffffff",
    "stripe": "#1a2030", "grad1": "#1f6feb", "grad2": "#8b5cf6",
}
F = {
    "title": ("Segoe UI", 16, "bold"), "h2": ("Segoe UI", 12, "bold"),
    "body": ("Segoe UI", 10), "sm": ("Segoe UI", 9),
    "mono": ("Consolas", 9), "btn": ("Segoe UI", 10, "bold"),
    "stat": ("Segoe UI", 24, "bold"), "statlbl": ("Segoe UI", 9),
}

# ============================================================
# BRIDGE
# ============================================================
class ZaloBridge:
    def __init__(self, node_path, bridge_path, on_event=None):
        self.node_path = node_path
        self.bridge_path = bridge_path
        self.on_event = on_event
        self.proc = None
        self.mid = 0
        self.pending = {}
        self.running = False

    def start(self):
        try:
            self.proc = subprocess.Popen(
                [self.node_path, self.bridge_path],
                stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.PIPE,
                text=True, encoding="utf-8", bufsize=1,
                creationflags=subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0,
            )
            self.running = True
            threading.Thread(target=self._read, daemon=True).start()
            return True
        except Exception as e:
            print(f"[Bridge] {e}")
            return False

    def stop(self):
        self.running = False
        if self.proc:
            try: self.proc.terminate(); self.proc.wait(3)
            except: pass

    def _read(self):
        while self.running and self.proc and self.proc.poll() is None:
            try:
                line = self.proc.stdout.readline().strip()
                if not line: continue
                msg = json.loads(line)
                mid, data, err = msg.get("id"), msg.get("data"), msg.get("error")
                
                # If mid is a known callback, call it but don't pop if it's progress
                if mid in self.pending:
                    # Logic: if data has 'current', 'page', 'action', etc., it's progress, don't pop
                    is_progress = isinstance(data, dict) and ("current" in data or "fetched" in data or "event" in data or "action" in data)
                    if is_progress:
                        self.pending[mid](data, err)
                    else:
                        self.pending.pop(mid)(data, err)
                elif mid == "ready":
                    if self.on_event: self.on_event("ready", data)
                elif self.on_event:
                    # Global events or status
                    self.on_event(mid, data)
            except json.JSONDecodeError: continue
            except: break

    def send(self, action, params=None, callback=None):
        if not self.proc or self.proc.poll() is not None:
            if callback: callback(None, "Bridge stopped")
            return
        self.mid += 1
        m = str(self.mid)
        if callback: self.pending[m] = callback
        try:
            self.proc.stdin.write(json.dumps({"id": m, "action": action, "params": params or {}}) + "\n")
            self.proc.stdin.flush()
        except Exception as e:
            if callback: callback(None, str(e))


# ============================================================
# WIDGETS
# ============================================================
class GradientCanvas(tk.Canvas):
    def __init__(self, master, c1, c2, h=50, **kw):
        super().__init__(master, height=h, highlightthickness=0, **kw)
        self.c1, self.c2 = c1, c2
        self.bind("<Configure>", self._draw)

    def _draw(self, e=None):
        self.delete("g")
        w, h = self.winfo_width(), self.winfo_height()
        r1,g1,b1 = int(self.c1[1:3],16), int(self.c1[3:5],16), int(self.c1[5:7],16)
        r2,g2,b2 = int(self.c2[1:3],16), int(self.c2[3:5],16), int(self.c2[5:7],16)
        for i in range(max(w,1)):
            t = i/max(w,1)
            c = f"#{int(r1+(r2-r1)*t):02x}{int(g1+(g2-g1)*t):02x}{int(b1+(b2-b1)*t):02x}"
            self.create_line(i,0,i,h,fill=c,tags="g")


class Btn(tk.Frame):
    def __init__(self, master, text, cmd=None, color=C["accent"], fg="white", **kw):
        fnt = kw.pop("font", F["btn"])
        bg = master.cget("bg") if hasattr(master,'cget') else C["bg"]
        super().__init__(master, bg=bg, **kw)
        self.color, self.cmd = color, cmd
        r,g,b = int(color[1:3],16), int(color[3:5],16), int(color[5:7],16)
        self.hover = f"#{min(255,r+30):02x}{min(255,g+30):02x}{min(255,b+30):02x}"
        self.lbl = tk.Label(self, text=text, font=fnt, bg=color, fg=fg,
                            padx=16, pady=6, cursor="hand2")
        self.lbl.pack(fill="x")
        self.lbl.bind("<Enter>", lambda e: self.lbl.config(bg=self.hover))
        self.lbl.bind("<Leave>", lambda e: self.lbl.config(bg=self.color))
        self.lbl.bind("<Button-1>", lambda e: self.cmd() if self.cmd else None)

    def set_enabled(self, ok):
        if ok:
            self.lbl.config(bg=self.color, cursor="hand2")
            self.lbl.bind("<Button-1>", lambda e: self.cmd() if self.cmd else None)
        else:
            self.lbl.config(bg=C["hover"], cursor="")
            self.lbl.unbind("<Button-1>")


def make_entry(parent, **kw):
    return tk.Entry(parent, font=F["mono"], bg=C["input"], fg=C["text"],
                    insertbackground=C["text"], highlightbackground=C["border"],
                    highlightthickness=1, relief="flat", bd=2, **kw)


def make_label(parent, text, **kw):
    return tk.Label(parent, text=text, font=F["sm"], bg=C["card"], fg=C["dim"], **kw)


def make_card(parent, title, color=C["accent"]):
    f = tk.LabelFrame(parent, text=f"  {title}  ", font=F["h2"],
                      bg=C["card"], fg=color, highlightbackground=C["border"],
                      highlightthickness=1, padx=10, pady=8, labelanchor="n")
    return f


# ============================================================
# MAIN APP
# ============================================================
class App:
    def __init__(self, root):
        self.root = root
        root.title("Zalo Group Scanner v2.1")
        root.geometry("1200x950")
        root.minsize(1050, 850)
        root.configure(bg=C["bg"])

        self.bridge = None
        self.logged_in = False
        self.scan_result = None
        self.filtered = []
        self.filter_admin = tk.BooleanVar(value=True)
        self.accounts = self._load_accounts()
        self.current_account = None
        self.batch_running = False
        self.image_path = tk.StringVar(value="Chưa chọn ảnh...")
        self.use_random = tk.BooleanVar(value=True)
        self.scan_history = self._load_scan_history()
        self.templates = self._load_templates()
        self.current_data_file = None # Theo dõi file JSON hiện tại để lưu đè status

        EXCEL_DATA_DIR.mkdir(exist_ok=True)
        self.node_path = self._find_node()
        self._build_ui()
        self._start_bridge()
        
        # Restore last session
        self.root.after(1000, self._load_last_session)

    # ---- Node.js ----
    def _find_node(self):
        p = BASE_DIR / "nodejs_portable" / "node-v20.12.2-win-x64" / "node.exe"
        if p.exists(): return str(p)
        import shutil
        return shutil.which("node")

    # ---- Accounts ----
    def _load_accounts(self):
        try:
            if CREDENTIALS_FILE.exists():
                d = json.loads(CREDENTIALS_FILE.read_text("utf-8"))
                accs = []
                if isinstance(d, list): accs = d
                elif isinstance(d, dict):
                    if "accounts" in d: accs = d["accounts"]
                    else: accs = [d] # Single account format
                
                # Ensure all accounts have a 'name'
                for i, a in enumerate(accs):
                    if "name" not in a: a["name"] = f"Acc {i+1}"
                return accs
        except: pass
        return []

    def _save_accounts(self):
        CREDENTIALS_FILE.write_text(
            json.dumps({"accounts": self.accounts}, ensure_ascii=False, indent=2), "utf-8")

    # ---- Scan Data Persistence ----
    def _load_scan_history(self):
        """Load danh sách file Excel đã lưu"""
        history = []
        if EXCEL_DATA_DIR.exists():
            for f in sorted(EXCEL_DATA_DIR.glob("*.xlsx"), key=lambda x: x.stat().st_mtime, reverse=True):
                history.append({"file": f, "name": f.name})
        return history

    def _save_scan_data(self, scan_result=None, auto=False):
        """Lưu kết quả quét trực tiếp vào Excel có kèm Avatar (Hỗ trợ Auto-update)"""
        if not HAS_OPENPYXL: return None
        
        # Nếu không truyền scan_result, lấy từ bộ nhớ hiện tại
        res = scan_result or self.scan_result
        if not res: return None
        
        g = res.get("groupInfo", {})
        mems = res.get("members", [])
        
        # Luôn dùng duy nhất 1 file cho mỗi nhóm (Tên nhóm + ID nhóm)
        g_name = self._safe_name(g.get("name","group"))
        g_id = self._safe_name(g.get("groupId","unknown"))
        fp = EXCEL_DATA_DIR / f"{g_name}_{g_id}.xlsx"
        
        # Nếu đang ở chế độ auto-update và file đang mở hợp lệ, giữ nguyên đường dẫn đó
        if auto and self.current_data_file and self.current_data_file.exists():
            fp = self.current_data_file
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Thành viên"
        
        # Style
        hdr_font = Font(bold=True, color="FFFFFF")
        hdr_fill = PatternFill(start_color="1F6FEB", end_color="1F6FEB", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Meta Data
        ws.append([f"Nhóm: {g.get('name')}", f"ID: {g.get('groupId')}", f"Tổng: {g.get('totalMember')}"])
        ws.append(["Quét lúc:", datetime.now().isoformat()])
        ws.append([])

        # Table Header: STT, Avatar, ID, Tên hiển thị, Tên Zalo, Vai trò, Kết bạn, Mời nhóm, Nhắn tin
        headers = ["STT", "Avatar", "Zalo ID", "Tên hiển thị", "Tên Zalo", "Vai trò", "Kết bạn", "Mời nhóm", "Nhắn tin"]
        ws.append(headers)
        for cell in ws[4]:
            cell.font, cell.fill, cell.alignment, cell.border = hdr_font, hdr_fill, center_align, border

        # Data rows
        for i, m in enumerate(mems):
            rid = i + 5
            ws.row_dimensions[rid].height = 45 
            
            f_stat = "v" if m.get("friend_sent") else ""
            i_stat = "v" if m.get("invite_sent") else ""
            m_stat = "v" if m.get("message_sent") else ""
            
            row_vals = [i+1, "", str(m.get("id")), m.get("dName"), m.get("zaloName"), m.get("role") or "Thành viên", f_stat, i_stat, m_stat]
            ws.append(row_vals)
            
            # Download Avatar (Chỉ làm khi tạo file mới hoặc file chưa có ảnh để tránh lag)
            if not auto:
                avt_url = m.get("avatar")
                if HAS_PILLOW and HAS_REQUESTS and avt_url:
                    try:
                        resp = requests.get(avt_url, timeout=3)
                        if resp.status_code == 200:
                            img_data = BytesIO(resp.content)
                            img = PILImage.open(img_data)
                            img.thumbnail((55, 55))
                            temp_p = AVATAR_CACHE_DIR / f"{m.get('id')}.png"
                            img.save(temp_p)
                            xl_img = XLImage(temp_p)
                            ws.add_image(xl_img, f"B{rid}")
                    except: pass
            
            for col in range(1, 10): # Update border for all 9 columns
                cell = ws.cell(row=rid, column=col)
                cell.border = border
                cell.alignment = Alignment(vertical="center", horizontal="center" if col in [1,2,7,8,9] else "left")

        # Column widths
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 28
        ws.column_dimensions["E"].width = 24
        
        try:
            wb.save(fp)
            self.current_data_file = fp
            # Update history memory
            if not any(h["file"] == fp for h in self.scan_history):
                self.scan_history = self._load_scan_history()
                self._refresh_history_combo()
        except Exception as e:
            print(f"Error saving excel: {e}")
            
        return fp

    def _refresh_history_combo(self):
        names = [h['name'] for h in self.scan_history]
        self.history_combo["values"] = names
        if names: self.history_combo.current(0)

    def _save_last_session(self):
        """Lưu toàn bộ thông số và data phiên làm việc"""
        try:
            data = {
                "link": self.link_entry.get().strip(),
                "delay": self.delay_entry.get().strip(),
                "limit": self.limit_entry.get().strip(),
                "target_group_id": self.group_id_entry.get().strip(),
                "use_random": self.use_random.get(),
                "scan_result": self.scan_result
            }
            LAST_SESSION_FILE.write_text(json.dumps(data, ensure_ascii=False), "utf-8")
            
            # Đồng bộ luôn vào file Excel hiện tại nếu có
            if self.current_data_file and self.current_data_file.exists() and self.scan_result:
                # Với Excel, ta sẽ ghi đè toàn bộ hoặc cập nhật status. 
                # Để đảm bảo đồng bộ, ta gọi lại hàm save Excel
                self._save_scan_data(self.scan_result)
        except: pass

    def _load_last_session(self):
        """Khôi phục toàn bộ thông số từ file"""
        try:
            if LAST_SESSION_FILE.exists():
                d = json.loads(LAST_SESSION_FILE.read_text("utf-8"))
                
                # Restore inputs
                if d.get("link"):
                    self.link_entry.delete(0, "end")
                    self.link_entry.insert(0, d["link"])
                if d.get("delay"):
                    self.delay_entry.delete(0, "end")
                    self.delay_entry.insert(0, d["delay"])
                if d.get("limit"):
                    self.limit_entry.delete(0, "end")
                    self.limit_entry.insert(0, d["limit"])

                if d.get("target_group_id"):
                    self.group_id_entry.delete(0, "end")
                    self.group_id_entry.insert(0, d["target_group_id"])
                if "use_random" in d:
                    self.use_random.set(d["use_random"])

                # Restore scan result and member list
                sr = d.get("scan_result")
                if sr and sr.get("members"):
                    self.scan_result = sr
                    self._apply_filter()
                    n = len(sr.get("members", []))
                    gname = sr.get("groupInfo", {}).get("name", "N/A")
                    self.plabel.config(text=f"✅ Đã tải: {n} mms từ {gname}")
                    self.pvar.set(100)
                    self._log(f"🔄 Đã khôi phục phiên trước: {n} mms — {gname}", "info")
                else:
                    self._log("🔄 Đã khôi phục cài đặt phiên trước.", "info")
        except Exception as e:
            self._log(f"⚠️ Lỗi khôi phục phiên: {e}", "warn")

    # ---- UI ----
    def _build_ui(self):
        # Create Notebook for Tabs
        self.nb = ttk.Notebook(self.root)
        self.nb.pack(fill="both", expand=True, padx=5, pady=5)

        # TAB 1: OPERATE
        self.tab_op = tk.Frame(self.nb, bg=C["bg"])
        self.nb.add(self.tab_op, text=" ⚡ VẬN HÀNH & QUÉT ")
        
        top_main = tk.Frame(self.tab_op, bg=C["bg"])
        top_main.pack(fill="both", expand=True)

        # LEFT COLUMN (340px)
        left = tk.Frame(top_main, bg=C["bg"], width=350)
        left.pack(side="left", fill="y", padx=(10, 5), pady=10)
        left.pack_propagate(False)

        self._build_account_section(left)
        self._build_scan_section(left)
        self._build_filter_section(left)
        self._build_action_section(left)

        # RIGHT COLUMN (Resizable Table & Logs)
        right = tk.Frame(top_main, bg=C["bg"])
        right.pack(side="left", fill="both", expand=True, padx=(5, 10), pady=10)
        self._build_resizable_panel(right)

        # TAB 2: TEMPLATES
        self.tab_tpl = tk.Frame(self.nb, bg=C["bg"])
        self.nb.add(self.tab_tpl, text=" 📜 THƯ VIỆN KỊCH BẢN ")
        self._build_template_tab(self.tab_tpl)

    # ---- ACCOUNT SECTION ----
    def _build_account_section(self, parent):
        card = make_card(parent, "🔐 TÀI KHOẢN", C["accent"])
        card.pack(fill="x", pady=(0,5))

        self.login_status = tk.Label(card, text="⏳ Đang chờ...", font=F["sm"],
                                     bg=C["card"], fg=C["yellow"])
        self.login_status.pack(fill="x")

        # Account dropdown
        row_acc = tk.Frame(card, bg=C["card"])
        row_acc.pack(fill="x", pady=2)
        tk.Label(row_acc, text="Acc:", font=F["sm"], bg=C["card"], fg=C["dim"]).pack(side="left")
        self.acc_var = tk.StringVar()
        self.acc_combo = ttk.Combobox(row_acc, textvariable=self.acc_var, state="readonly", font=F["mono"])
        self.acc_combo.pack(side="left", fill="x", expand=True, padx=(5,0))
        self._refresh_account_list()

        btn_row = tk.Frame(card, bg=C["card"])
        btn_row.pack(fill="x", pady=2)
        Btn(btn_row, "🔑 Login", cmd=self._do_login, color=C["accent"]).pack(side="left", expand=True, fill="x", padx=(0,2))
        Btn(btn_row, "➕ Thêm", cmd=self._add_account, color="#2d6a4f").pack(side="left", expand=True, fill="x", padx=(1,1))
        Btn(btn_row, "🗑️ Xóa", cmd=self._delete_account, color=C["red"]).pack(side="left", expand=True, fill="x", padx=(2,0))

    def _refresh_account_list(self):
        names = [a.get("name", a.get("imei", "?")[:20]) for a in self.accounts]
        self.acc_combo["values"] = names
        if names:
            self.acc_combo.current(0)

    def _add_account(self):
        win = tk.Toplevel(self.root)
        win.title("Thêm tài khoản")
        win.geometry("520x480")
        win.configure(bg=C["bg"])
        win.transient(self.root)
        win.grab_set()

        card = tk.Frame(win, bg=C["card"], padx=16, pady=16)
        card.pack(fill="both", expand=True, padx=16, pady=16)

        tk.Label(card, text="Tên tài khoản:", font=F["body"], bg=C["card"],
                 fg=C["text"]).pack(anchor="w")
        name_e = make_entry(card)
        name_e.pack(fill="x", pady=(2,8))

        tk.Label(card, text="IMEI:", font=F["body"], bg=C["card"],
                 fg=C["text"]).pack(anchor="w")
        imei_e = make_entry(card)
        imei_e.pack(fill="x", pady=(2,8))

        tk.Label(card, text="Cookie (JSON array từ ZaloDataExtractor):", font=F["body"],
                 bg=C["card"], fg=C["text"]).pack(anchor="w")
        cookie_t = tk.Text(card, font=F["mono"], bg=C["input"], fg=C["text"],
                           insertbackground=C["text"], height=8, relief="flat", bd=5, wrap="word")
        cookie_t.pack(fill="x", pady=(2,8))

        tk.Label(card, text="User-Agent:", font=F["body"], bg=C["card"],
                 fg=C["text"]).pack(anchor="w")
        ua_e = make_entry(card)
        ua_e.insert(0, DEFAULT_USER_AGENT)
        ua_e.pack(fill="x", pady=(2,8))

        def save():
            name = name_e.get().strip() or f"Account {len(self.accounts)+1}"
            imei = imei_e.get().strip()
            raw = cookie_t.get("1.0", "end").strip()
            ua = ua_e.get().strip() or DEFAULT_USER_AGENT
            if not imei or not raw:
                messagebox.showwarning("Thiếu", "Cần IMEI và Cookie!", parent=win)
                return
            try:
                cookie = json.loads(raw)
            except:
                messagebox.showerror("Lỗi", "Cookie phải là JSON hợp lệ!", parent=win)
                return
            self.accounts.append({"name": name, "imei": imei, "cookie": cookie, "userAgent": ua})
            self._save_accounts()
            self._refresh_account_list()
            self.acc_combo.current(len(self.accounts) - 1)
            self._log(f"✅ Thêm tài khoản: {name}", "ok")
            win.destroy()

        Btn(card, "💾 Lưu tài khoản", cmd=save, color=C["green"]).pack(fill="x", pady=(4,0))

    def _delete_account(self):
        idx = self.acc_combo.current()
        if idx < 0: return
        name = self.accounts[idx].get("name", "?")
        if not messagebox.askyesno("Xác nhận", f"Xóa tài khoản '{name}'?"):
            return
        self.accounts.pop(idx)
        self._save_accounts()
        self._refresh_account_list()
        self._log(f"🗑️ Đã xóa tài khoản: {name}", "warn")

    # ---- SCAN SECTION ----
    def _build_scan_section(self, parent):
        card = make_card(parent, "🔍 QUÉT NHÓM", C["green"])
        card.pack(fill="x", pady=(0,5))

        make_label(card, "Link nhóm Zalo:").pack(anchor="w")
        self.link_entry = make_entry(card)
        self.link_entry.pack(fill="x", pady=(2,2))
        self.link_entry.insert(0, "https://zalo.me/g/")

        # Tiến trình (ẩn đi theo yêu cầu xóa ô thừa)
        self.pvar = tk.DoubleVar()
        self.plabel = tk.Label(card, text="", font=F["sm"], bg=C["card"], fg=C["dim"])

        self.scan_btn = Btn(card, "🚀 BẮT ĐẦU QUÉT", cmd=self._do_scan, color=C["green"])
        self.scan_btn.pack(fill="x", pady=(8,0))

    # ---- FILTER ----
    def _build_filter_section(self, parent):
        card = make_card(parent, "⚙️ BỘ LỌC · XUẤT FILE", C["purple"])
        card.pack(fill="x", pady=(0,5))

        tk.Checkbutton(card, text=" Loại bỏ Trưởng/Phó nhóm", variable=self.filter_admin,
                       font=F["body"], bg=C["card"], fg=C["text"], selectcolor=C["input"],
                       activebackground=C["card"], activeforeground=C["text"],
                       command=self._apply_filter).pack(anchor="w")

        # --- ROW: EXCEL EXPORT & IMPORT ---
        row_ex = tk.Frame(card, bg=C["card"])
        row_ex.pack(fill="x", pady=(4,0))
        
        Btn(row_ex, "📗 Xuất Excel", cmd=self._export_excel, color="#065f46"
            ).pack(side="left", expand=True, fill="x", padx=(0,2))
            
        Btn(row_ex, "📂 Import Excel", cmd=self._import_excel, color="#374151"
            ).pack(side="left", expand=True, fill="x", padx=(2,0))

    # ---- ACTION SECTION ----
    def _build_action_section(self, parent):
        card = make_card(parent, "🚀 HÀNH ĐỘNG HÀNG LOẠT", C["orange"])
        card.pack(fill="x", pady=(0,5))

        # Row: Delay + Limit
        row_dl = tk.Frame(card, bg=C["card"])
        row_dl.pack(fill="x", pady=(0,4))

        dl_f = tk.Frame(row_dl, bg=C["card"])
        dl_f.pack(side="left", expand=True, fill="x", padx=(0,4))
        make_label(dl_f, "Delay (giây):").pack(anchor="w")
        self.delay_entry = make_entry(dl_f, width=8)
        self.delay_entry.insert(0, "30")
        self.delay_entry.pack(fill="x")

        lm_f = tk.Frame(row_dl, bg=C["card"])
        lm_f.pack(side="left", expand=True, fill="x", padx=(4,0))
        make_label(lm_f, "Giới hạn (0=tất cả):").pack(anchor="w")
        self.limit_entry = make_entry(lm_f, width=8)
        self.limit_entry.insert(0, "0")
        self.limit_entry.pack(fill="x")

        # --- KỊCH BẢN ---
        make_label(card, "👋 Chọn kịch bản Kết bạn:").pack(anchor="w", pady=(4,0))
        self.tpl_friend_sel = tk.StringVar()
        self.tpl_friend_combo = ttk.Combobox(card, textvariable=self.tpl_friend_sel, state="readonly", font=F["body"])
        self.tpl_friend_combo.pack(fill="x", pady=(2,4))

        make_label(card, "📨 Chọn kịch bản Mời nhóm:").pack(anchor="w", pady=(4,0))
        self.tpl_invite_sel = tk.StringVar()
        self.tpl_invite_combo = ttk.Combobox(card, textvariable=self.tpl_invite_sel, state="readonly", font=F["body"])
        self.tpl_invite_combo.pack(fill="x", pady=(2,4))

        make_label(card, "💬 Chọn kịch bản Nhắn tin:").pack(anchor="w", pady=(4,0))
        self.tpl_msg_sel = tk.StringVar()
        self.tpl_msg_combo = ttk.Combobox(card, textvariable=self.tpl_msg_sel, state="readonly", font=F["body"])
        self.tpl_msg_combo.pack(fill="x", pady=(2,4))
        
        # Random Mode Toggle
        tk.Checkbutton(card, text=" Gửi nội dung Ngẫu nhiên (Ưu tiên)", 
                       variable=self.use_random, font=F["sm"], bg=C["card"], fg=C["orange"],
                       selectcolor=C["input"], activebackground=C["card"]).pack(anchor="w", pady=2)

        # --- ĐỊNH DANH NHÓM ---
        make_label(card, "🆔 ID nhóm mời (mời trực tiếp):").pack(anchor="w")
        self.group_id_entry = make_entry(card)
        self.group_id_entry.pack(fill="x", pady=(2,4))
        
        # Biến ẩn để chứa ảnh từ kịch bản (không hiển thị UI chọn ảnh ở đây nữa)
        self.current_action_image = tk.StringVar(value="")

        self.action_progress = tk.Label(card, text="", font=F["sm"], bg=C["card"], fg=C["dim"])
        self.action_progress.pack(fill="x", pady=(2,2))

        # --- NÚT BẤM ---
        r1 = tk.Frame(card, bg=C["card"])
        r1.pack(fill="x", pady=(4,2))
        Btn(r1, "👋 Kết bạn", cmd=self._batch_friend, color=C["accent"]).pack(side="left", expand=True, fill="x", padx=(0,3))
        Btn(r1, "📨 Mời nhóm", cmd=self._batch_invite, color=C["orange"]).pack(side="left", expand=True, fill="x", padx=(3,0))

        Btn(card, "💥 GỬI TIN NHẮN (Marketing)", cmd=self._batch_message, color=C["green"]).pack(fill="x", pady=(2,2))
        self.stop_btn = Btn(card, "⛔ DỪNG", cmd=self._do_cancel, color=C["red"])
        self.stop_btn.pack(fill="x", pady=(2,0))

        self._refresh_action_combos()
        self.stop_btn.pack(fill="x", pady=(2,0))

        # ---- Scan History ----
        sep = tk.Frame(card, bg=C["border"], height=1)
        sep.pack(fill="x", pady=(8,4))
        make_label(card, "📂 Lịch sử quét (load lại):").pack(anchor="w")
        self.history_var = tk.StringVar()
        self.history_combo = ttk.Combobox(card, textvariable=self.history_var,
                                          state="readonly", font=F["mono"])
        self.history_combo.pack(fill="x", pady=(2,2))
        self._refresh_history()
        Btn(card, "📗 Load Excel đã lưu", cmd=self._load_history_item,
            color="#374151").pack(fill="x", pady=(2,0))

    # ---- STATS ----
    def _build_stats(self, parent):
        row = tk.Frame(parent, bg=C["bg"])
        row.pack(fill="x", pady=(0,4))

        def stat(parent, icon, lbl, color):
            f = tk.Frame(parent, bg=C["card2"], highlightbackground=C["border"],
                         highlightthickness=1, padx=8, pady=4)
            tk.Label(f, text=icon, font=F["body"], bg=C["card2"], fg=color).pack(side="left")
            v = tk.Label(f, text="0", font=("Segoe UI",11,"bold"), bg=C["card2"], fg=C["bright"])
            v.pack(side="left", padx=5)
            tk.Label(f, text=lbl, font=("Segoe UI",8), bg=C["card2"], fg=C["dim"]).pack(side="left")
            return f, v

        self.s_total_f, self.s_total = stat(row, "👥", "Tổng", C["accent"])
        self.s_total_f.pack(side="left", fill="x", expand=True, padx=(0,2))
        self.s_admin_f, self.s_admin = stat(row, "👑", "Admin", C["yellow"])
        self.s_admin_f.pack(side="left", fill="x", expand=True, padx=2)
        self.s_member_f, self.s_member = stat(row, "✅", "Mems", C["green"])
        self.s_member_f.pack(side="left", fill="x", expand=True, padx=(2,0))


    # ---- TABLE & LOGS (RESIZABLE) ----
    def _build_resizable_panel(self, parent):
        self.pane = tk.PanedWindow(parent, orient="vertical", bg=C["bg"], sashwidth=4, sashrelief="flat")
        self.pane.pack(fill="both", expand=True)

        # Upper: Table Area
        tf = tk.Frame(self.pane, bg=C["bg"])
        self.pane.add(tf, minsize=300)
        
        # Add stats first
        self._build_stats(tf)

        # Table Frame
        tbl_f = tk.Frame(tf, bg=C["card"], highlightbackground=C["border"], highlightthickness=1)
        tbl_f.pack(fill="both", expand=True)

        hdr = tk.Frame(tbl_f, bg=C["card2"], pady=5, padx=8)
        hdr.pack(fill="x")
        self.tbl_title = tk.Label(hdr, text="📋 DANH SÁCH THÀNH VIÊN", font=F["h2"], bg=C["card2"], fg=C["text"])
        self.tbl_title.pack(side="left")
        
        sel_row = tk.Frame(hdr, bg=C["card2"])
        sel_row.pack(side="right")
        Btn(sel_row, "✅ Tất cả", cmd=lambda: self._toggle_all(True), color=C["accent"], font=F["sm"]).pack(side="left", padx=5)
        Btn(sel_row, "⬜ Bỏ chọn", cmd=lambda: self._toggle_all(False), color="#374151", font=F["sm"]).pack(side="left")

        tc = tk.Frame(tbl_f, bg=C["card"])
        tc.pack(fill="both", expand=True, padx=1)
        
        cols = ("check", "stt", "id", "name", "zname", "friend", "invite", "msg", "status")
        self.tree = ttk.Treeview(tc, columns=cols, show="headings", style="T.Treeview", selectmode="extended")
        
        col_defs = [
            ("check", "Chon", 35, "center"),
            ("stt", "STT", 40, "center"),
            ("id", "Zalo ID", 155, "w"),
            ("name", "Tên hiển thị", 210, "w"),
            ("zname", "Tên Zalo", 160, "w"),
            ("friend", "Kết bạn", 80, "center"),
            ("invite", "Mời nhóm", 80, "center"),
            ("msg", "Nhắn tin", 80, "center"),
            ("status", "Kết quả", 120, "center")
        ]
        for cid, head, w, anc in col_defs:
            self.tree.heading(cid, text=head if cid != "check" else "☑")
            self.tree.column(cid, width=w, anchor=anc)

        self.tree.tag_configure("ok", foreground=C["green"])
        self.tree.tag_configure("warn", foreground=C["yellow"])
        self.tree.tag_configure("error", foreground=C["red"])
        self.tree.bind("<Button-1>", self._on_tree_click)
        
        self.tree.pack(side="left", fill="both", expand=True)
        sb = ttk.Scrollbar(tc, orient="vertical", command=self.tree.yview)
        sb.pack(side="right", fill="y")
        self.tree.configure(yscrollcommand=sb.set)

        # Lower: Logs
        lf = tk.Frame(self.pane, bg=C["card"], highlightbackground=C["border"], highlightthickness=1)
        self.pane.add(lf, minsize=100)

        tk.Frame(lf, bg=C["card2"], pady=3, padx=8).pack(fill="x")
        tk.Label(lf, text="📝 NHẬT KÝ HOẠT ĐỘNG (Có thể kéo dãn)", font=F["body"], bg=C["card2"], fg=C["dim"]).place(x=8, y=2)
        
        self.log = tk.Text(lf, font=F["mono"], bg=C["card"], fg=C["dim"], height=4,
                           state="disabled", relief="flat", bd=6, wrap="word")
        self.log.pack(fill="both", expand=True, pady=(20,0))
        for t,c in [("info",C["accent"]),("ok",C["green"]),("warn",C["yellow"]),("error",C["red"])]:
            self.log.tag_configure(t, foreground=c)

    def _log(self, msg, tag="info", acc_name=None):
        ts = datetime.now().strftime("%H:%M:%S")
        prefix = f"[Acc:{acc_name}] " if acc_name else ""
        self.log.config(state="normal")
        self.log.insert("end", f"[{ts}] {prefix}{msg}\n", tag)
        self.log.see("end")
        self.log.config(state="disabled")

    # ============================================================
    # BRIDGE
    # ============================================================
    def _start_bridge(self):
        if not self.node_path:
            self.login_status.config(text="❌ Node.js not found!", fg=C["red"])
            return
        bp = str(BASE_DIR / "zalo_bridge.mjs")
        self.bridge = ZaloBridge(self.node_path, bp, on_event=lambda e,d: self.root.after(0, self._evt, e, d))
        if self.bridge.start():
            self._log("Bridge khởi động...")
        else:
            self.login_status.config(text="❌ Bridge failed", fg=C["red"])

    def _evt(self, eid, d):
        if eid == "ready":
            self.login_status.config(text="🔵 Bridge sẵn sàng", fg=C["accent"])
            self._log("Hệ thống Bridge đã hoạt động.", "ok")
            if self.accounts:
                self.root.after(300, self._do_login)
        elif eid == "scan_progress":
            fetched = d.get('fetched', 0)
            self.plabel.config(text=f"Đang quét trang {d.get('page')}... ({fetched} mems)")
            # Indeterminate progress or based on a guess
            self.pvar.set((self.pvar.get() + 5) % 100)

    def _batch_progress_handler(self, d, err, acc_name, act):
        if err:
            self.root.after(0, self._log, f"❌ Lỗi {act}: {err}", "error", acc_name)
            return
        if not d: return
        
        # Check message type: progress (has 'current'), debug (has 'action' but no 'current'), or final (has 'success')
        if "current" in d:
            target_id = str(d.get("uid") or "")
            ok = d.get("ok", False)
            curr = d.get("current", 0)
            total = d.get("total", 0)
            
            # Chỉ cập nhật trạng thái "đã gửi" khi THÀNH CÔNG
            members = self.scan_result.get("members", [])
            found_member = None
            for m in members:
                if str(m.get("id")) == target_id:
                    if ok:  # CHỈ đánh dấu khi thực sự thành công
                        if act == "Kết bạn": m["friend_sent"] = True
                        elif act == "Mời nhóm": m["invite_sent"] = True
                        elif act == "Nhắn tin": m["message_sent"] = True
                    found_member = m
                    break
            
            # Cập nhật UI ngay lập tức cho dòng đó
            if found_member:
                for item_id in self.tree.get_children():
                    if str(self.tree.item(item_id, "values")[2]) == target_id:
                        f_stat = "✅" if found_member.get("friend_sent") else ""
                        i_stat = "✅" if found_member.get("invite_sent") else ""
                        m_stat = "✅" if found_member.get("message_sent") else ""
                        
                        old_vals = list(self.tree.item(item_id, "values"))
                        old_vals[5] = f_stat
                        old_vals[6] = i_stat
                        old_vals[7] = m_stat
                        tag = "ok" if ok else "error"
                        self.tree.item(item_id, values=old_vals, tags=(tag,))
                        break
            
            # Tự động lưu vào Excel để bảo toàn dữ liệu
            if ok:
                self._save_scan_data(auto=True)
            
            # Log with Account Name
            if ok:
                self._log(f"✅ {act} {target_id} ({curr}/{total})", "ok", acc_name)
            else:
                self._log(f"❌ {act} {target_id} lỗi: {d.get('error','?')}", "warn", acc_name)
            
            # Update progress label
            self.action_progress.config(text=f"{act}: {curr}/{total}")
            
            # Update Table View
            self.root.after(0, self._update_row_status, target_id, act, ok)
        elif "action" in d and "current" not in d:
            # Debug/info message from bridge (e.g. invite_debug) - just log, don't say "done"
            self._log(f"ℹ️ {act}: {d}", "info", acc_name)
        elif "success" in d:
            # Final result from the batch
            ok_count = d.get("successCount", 0)
            fail_count = d.get("failCount", 0)
            self._log(f"🏁 Xong {act} tài khoản {acc_name}: ✅{ok_count} ❌{fail_count}", "ok", acc_name)
            self.action_progress.config(text=f"Hoàn thành {act}")
            self._save_last_session()
        else:
            # Unknown format, log it
            self._log(f"🏁 Xong {act} tài khoản {acc_name}", "ok", acc_name)

    def _update_row_status(self, uid, act, ok):
        if not self.scan_result: return
        mems = self.scan_result.get("members", [])
        
        # Mapping hành động sang cột và thuộc tính
        # Index: 5:Kết bạn, 6:Mời nhóm, 7:Nhắn tin
        action_map = {
            "Kết bạn": {"idx": 5, "attr": "friend_sent"},
            "Mời nhóm": {"idx": 6, "attr": "invite_sent"},
            "Nhắn tin": {"idx": 7, "attr": "message_sent"}
        }
        
        cfg = action_map.get(act)
        if not cfg: return

        # Update bộ nhớ
        for m in mems:
            if str(m.get("id")) == str(uid):
                if ok: m[cfg["attr"]] = True
                break

        # Update Treeview
        for item_id in self.tree.get_children():
            vals = list(self.tree.item(item_id, "values"))
            if str(vals[2]) == str(uid):
                vals[cfg["idx"]] = "✅" if ok else "❌"
                vals[8] = "Thành công" if ok else "Lỗi"
                self.tree.item(item_id, values=vals, tags=("ok" if ok else "error"))
                break

    # ============================================================
    # LOGIN
    # ============================================================
    def _do_login(self):
        acc = self._get_current_acc()
        if not acc: return
        self.login_status.config(text="⏳ Đang đăng nhập...", fg=C["yellow"])
        self._log(f"Đăng nhập: {acc['name']}...", "info", acc['name'])

        def cb(data, err):
            if err: self._log(f"Lỗi: {err}", "error", acc['name'])
            elif data: 
                self.root.after(0, lambda: self.login_status.config(text=f"🟢 {acc['name']} OK", fg=C["green"]))
                self._log(f"Đăng nhập thành công!", "ok", acc['name'])

        self.bridge.send("login", {
            "imei": acc["imei"], "cookie": acc["cookie"],
            "userAgent": acc.get("userAgent", DEFAULT_USER_AGENT)
        }, cb)

    def _login_result(self, data, err):
        if err:
            self.login_status.config(text=f"❌ {err}", fg=C["red"])
            self._log(f"Đăng nhập thất bại: {err}", "error")
            return
        if data and data.get("success"):
            self.logged_in = True
            name = self.current_account.get("name","?") if self.current_account else "?"
            self.login_status.config(text=f"🟢 {name} — Đã đăng nhập!", fg=C["green"])
            self._log(f"✅ Đăng nhập thành công: {name}", "ok")
        else:
            msg = data.get("message","?") if data else "?"
            self.login_status.config(text=f"❌ {msg}", fg=C["red"])
            self._log(f"Thất bại: {msg}", "error")

    # = = = = = = = = = = = = = = = = = = = = = = = = = = = = = = = = = = = = = = = 
    # SCAN
    # = = = = = = = = = = = = = = = = = = = = = = = = = = = = = = = = = = = = = = = 
    def _do_scan(self):
        link = self.link_entry.get().strip()
        if "zalo.me/g/" not in link:
            return messagebox.showwarning("!", "Link phải có dạng https://zalo.me/g/xxxxx")
        
        acc = self._get_current_acc()
        if not acc: return messagebox.showwarning("!", "Chọn tài khoản trước!")

        acc_name = acc.get("name", "Acc")
        self._log(f"🔎 Bắt đầu quét: {link}", "info", acc_name)
        self.plabel.config(text="Đang quét...")

        def thread_task():
            params = {"link": link, "imei": acc['imei'], "cookie": acc['cookie']}
            def cb(data, err):
                if err: self._log(f"Lỗi quét: {err}", "error", acc_name)
                elif data and data.get("success"):
                    self.root.after(0, self._process_scan_result, data, acc_name)
            self.bridge.send("scan_group", params, cb)

        threading.Thread(target=thread_task, daemon=True).start()

    def _process_scan_result(self, data, acc_name):
        self.plabel.config(text="✅ Đã quét xong!")
        
        try:
            fp = self._save_scan_data(data)
            self._log(f"📂 Quét & Gộp xong: {len(self.scan_result['members'])} mms", "ok", acc_name)
            self._apply_filter()
            self._save_last_session()
        except Exception as e:
            self._log(f"Lỗi lưu data: {e}", "warn")

    def _refresh_history(self):
        self.scan_history = self._load_scan_history()
        names = []
        for h in self.scan_history:
            # Lấy thời gian sửa đổi file
            try:
                mtime = h["file"].stat().st_mtime
                ts = datetime.fromtimestamp(mtime).strftime("%Y-%m-%d %H:%M")
            except:
                ts = "N/A"
            names.append(f"{h['name']} ({ts})")
            
        self.history_combo["values"] = names
        if names: self.history_combo.current(0)

    def _load_history_item(self):
        idx = self.history_combo.current()
        if idx < 0: return
        item = self.scan_history[idx]
        self._import_excel_file(str(item["file"]))

    def _import_excel_file(self, fp):
        if not HAS_OPENPYXL:
            messagebox.showerror("!", "Cần cài openpyxl để load Excel.")
            return
        
        try:
            wb = load_workbook(fp, data_only=True)
            ws = wb.active
            
            # Read Group Meta from row 1
            g_name = str(ws.cell(row=1, column=1).value or "").replace("Nhóm:", "").strip()
            g_id = str(ws.cell(row=1, column=2).value or "").replace("ID:", "").strip()
            
            members = []
            header_row = 4 # Based on our new template
            
            for r in range(header_row + 1, ws.max_row + 1):
                mid = str(ws.cell(row=r, column=3).value or "").strip()
                if not mid or mid == "None": continue
                
                name = str(ws.cell(row=r, column=4).value or "").strip()
                zname = str(ws.cell(row=r, column=5).value or "").strip()
                role = str(ws.cell(row=r, column=6).value or "").strip()
                f_sent = str(ws.cell(row=r, column=7).value or "").lower() == "v"
                i_sent = str(ws.cell(row=r, column=8).value or "").lower() == "v"
                m_sent = str(ws.cell(row=r, column=9).value or "").lower() == "v"
                
                members.append({
                    "id": mid, "dName": name, "zaloName": zname, "role": role,
                    "friend_sent": f_sent, "invite_sent": i_sent, "message_sent": m_sent
                })

            self.scan_result = {
                "groupInfo": {"name": g_name, "groupId": g_id},
                "members": members
            }
            self.current_data_file = Path(fp)
            self._log(f"📥 Đã tải {len(members)} mms từ Excel: {self.current_data_file.name}", "ok")
            self._apply_filter()
            
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể đọc file Excel: {e}")

    def _do_cancel(self):
        if self.bridge:
            self.bridge.send("cancel")
            self._log("⛔ Đang yêu cầu dừng...", "warn")
            self.batch_running = False

    # ============================================================
    # FILTER
    # ============================================================
    def _apply_filter(self):
        if not self.scan_result: return
        g = self.scan_result.get("groupInfo",{})
        mems = self.scan_result.get("members",[])
        cid = g.get("creatorId","")
        aids = set(g.get("adminIds",[]))

        all_d = []
        ac = 0
        for m in mems:
            mid = str(m.get("id") or "")
            existing_role = m.get("role")
            
            if mid == cid:
                role, is_a = "👑 Trưởng nhóm", True
            elif mid in aids:
                role, is_a = "⭐ Phó nhóm", True
            elif existing_role and ("Trưởng" in str(existing_role) or "Phó" in str(existing_role)):
                role, is_a = existing_role, True
            else:
                role, is_a = existing_role or "Thành viên", False
                
            if is_a: ac += 1
            all_d.append({
                "id": mid, 
                "dName": m.get("dName", m.get("displayName","")),
                "zaloName": m.get("zaloName",""), 
                "role": role, 
                "is_admin": is_a,
                "friend_sent": m.get("friend_sent", False),
                "invite_sent": m.get("invite_sent", False),
                "message_sent": m.get("message_sent", False),
                "status": m.get("last_status", "Sẵn sàng"),
                "checked": m.get("checked", True)
            })

        self.filtered = [m for m in all_d if not m["is_admin"]] if self.filter_admin.get() else all_d
        rc = len(mems) - ac

        self.s_total.config(text=str(len(mems)))
        self.s_admin.config(text=str(ac))
        self.s_member.config(text=str(rc))

        self.tree.delete(*self.tree.get_children())
        for i, m in enumerate(self.filtered):
            tags = []
            if "Trưởng" in m["role"]: tags.append("owner")
            elif "Phó" in m["role"]: tags.append("admin")
            if i%2==1: tags.append("stripe")
            
            check_mark = "☑" if m["checked"] else "☐"
            f_mark = "✅" if m["friend_sent"] else ""
            i_mark = "✅" if m["invite_sent"] else ""
            m_mark = "✅" if m["message_sent"] else ""
            
            # vals: check, stt, id, name, zname, friend, invite, msg, status
            vals = (check_mark, i+1, m["id"], m["dName"], m["zaloName"], f_mark, i_mark, m_mark, m["status"])
            self.tree.insert("","end", values=vals, tags=tuple(tags))

        ft = "(đã lọc)" if self.filter_admin.get() else ""
        self.tbl_title.config(text=f"📋 {g.get('name','N/A')} {ft}")

    # ============================================================
    # EXPORT
    # ============================================================
    def _save_template(self, name, content, t_type):
        # ... logic lưu ...
        self._refresh_tpl_combos()
        self._refresh_tpl_table()

    def _delete_template(self, name):
        # ... logic xóa ...
        self._refresh_tpl_combos()
        self._refresh_tpl_table()

    def _safe_name(self, n):
        return re.sub(r'[<>:"/\\|?*]', '_', n or "group")[:80]

    def _get_save_path(self, ext):
        if not self.filtered:
            messagebox.showinfo("!", "Chưa có dữ liệu!")
            return None
        g = self.scan_result.get("groupInfo",{}) if self.scan_result else {}
        nm = self._safe_name(g.get("name","zalo"))
        return filedialog.asksaveasfilename(
            defaultextension=ext,
            initialfile=f"{nm}_{datetime.now().strftime('%Y%m%d_%H%M')}{ext}",
            filetypes=[(f"{ext.upper()} files", f"*{ext}"), ("All", "*.*")])

    def _export_excel(self):
        if not HAS_OPENPYXL:
            messagebox.showerror("!", "Cần cài openpyxl:\npip install openpyxl")
            return
        fp = self._get_save_path(".xlsx")
        if not fp: return

        g = self.scan_result.get("groupInfo",{}) if self.scan_result else {}
        wb = Workbook()
        ws = wb.active
        ws.title = "Thành viên"

        # Style
        hdr_font = Font(bold=True, color="FFFFFF", size=11)
        hdr_fill = PatternFill(start_color="1F6FEB", end_color="1F6FEB", fill_type="solid")
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Info rows
        ws.append([f"Nhóm: {g.get('name','')}"])
        ws.append([f"Tổng thành viên: {g.get('totalMember','')}"])
        ws.append([f"Quét lúc: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        ws.append([f"Lọc admin: {'Có' if self.filter_admin.get() else 'Không'}"])
        ws.append([])

        # Header
        headers = ["STT", "Zalo ID", "Tên hiển thị", "Tên Zalo", "Vai trò", "Kết bạn", "Mời nhóm"]
        ws.append(headers)
        for cell in ws[6]:
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        for i, m in enumerate(self.filtered):
            f_stat = "v" if m.get("friend_sent") else ""
            i_stat = "v" if m.get("invite_sent") else ""
            row = [i+1, m["id"], m["dName"], m["zaloName"], m["role"], f_stat, i_stat]
            ws.append(row)
            for cell in ws[ws.max_row]:
                cell.border = border

        # Column widths
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 28
        ws.column_dimensions["D"].width = 24
        ws.column_dimensions["E"].width = 16

        wb.save(fp)
        self._log(f"💾 Excel: {fp}", "ok")
        messagebox.showinfo("✅", f"Đã xuất {len(self.filtered)} thành viên!\n{fp}")

    def _import_excel(self):
        if not HAS_OPENPYXL:
            messagebox.showerror("!", "Cần cài openpyxl để load Excel.")
            return
        
        fp = filedialog.askopenfilename(
            title="Chọn file Excel",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if fp:
            self._import_excel_file(fp)
            self._save_last_session()

    # ============================================================
    # BATCH ACTIONS
    # ============================================================
    def _get_selected_ids(self):
        # 1. Nếu có hàng đang bôi đậm (selection) -> lấy theo selection
        sel = self.tree.selection()
        if sel:
            return [self.tree.item(s)["values"][2] for s in sel]
        
        # 2. Nếu không bôi đậm -> lấy tất cả những hàng đang hiện ☑
        ids = []
        for item_id in self.tree.get_children():
            vals = self.tree.item(item_id, "values")
            if vals[0] == "☑":
                ids.append(vals[2])
        return ids

    def _toggle_all(self, state):
        mark = "☑" if state else "☐"
        for item_id in self.tree.get_children():
            vals = list(self.tree.item(item_id, "values"))
            vals[0] = mark
            self.tree.item(item_id, values=vals)

    def _on_tree_click(self, event):
        region = self.tree.identify("region", event.x, event.y)
        if region == "heading": return
        
        item_id = self.tree.identify_row(event.y)
        column = self.tree.identify_column(event.x)
        
        if item_id and column == "#1": # Column check
            vals = list(self.tree.item(item_id, "values"))
            vals[0] = "☑" if vals[0] == "☐" else "☐"
            self.tree.item(item_id, values=vals)
            return "break" # Prevent selection change if clicking checkbox

    def _get_delay(self):
        try: return max(1, int(float(self.delay_entry.get().strip()))) * 1000
        except: return 5000

    def _get_current_acc(self):
        idx = self.acc_combo.current()
        if idx >= 0 and idx < len(self.accounts):
            return self.accounts[idx]
        return None

    def _get_ids_for_action(self, action_type=None):
        ids = self._get_selected_ids()
        acc = self._get_current_acc()
        if not acc or not ids: return None, None
        
        # Lọc bỏ những người đã được xử lý thành công trước đó
        m_map = {m["id"]: m for m in self.filtered}
        final_ids = []
        
        for uid in ids:
            m = m_map.get(str(uid), {})
            if action_type == "friend" and m.get("friend_sent"): continue
            if action_type == "invite" and m.get("invite_sent"): continue
            if action_type == "message" and m.get("message_sent"): continue
            final_ids.append(str(uid)) # Buộc phải là string để tránh lỗi Bridge làm tròn số
            
        if not final_ids:
            messagebox.showinfo("Thông báo", "Tất cả những người được chọn đã được xử lý trước đó!")
            return None, None
            
        # Kiểm tra nhanh nếu thấy ID có dấu hiệu bị lỗi làm tròn (kết thúc bằng 000)
        if any(str(uid).endswith("000") for uid in final_ids):
            self._log("⚠️ Cảnh báo: Phát hiện ID có thể bị sai (do phiên cũ). Nếu mời lỗi, hãy QUÉT LẠI nhóm này.", "warn")
            
        limit = self._get_limit()
        return final_ids[:limit] if limit > 0 else final_ids, acc

    def _get_limit(self):
        try: return max(0, int(self.limit_entry.get().strip()))
        except: return 0

    def _batch_friend(self):
        actual_ids, acc = self._get_ids_for_action("friend")
        if not actual_ids: return
        
        delay = self._get_delay()
        gid = self.scan_result.get("groupInfo",{}).get("groupId","") if self.scan_result else ""
        use_rand = self.use_random.get()
        sel_tpl = self.tpl_friend_sel.get()
        tpls = self.templates.copy()

        def run():
            msg = "Xin chào!"
            if use_rand:
                f_tpls = [t for t in tpls if t["type"] == "Kết bạn"]
                if f_tpls: msg = random.choice(f_tpls)["content"]
            else:
                for t in tpls:
                    if t.get("name") == sel_tpl: msg = t["content"]; break
            
            self._log(f"👋 Kết bạn: {len(actual_ids)} người", "info", acc['name'])
            self.bridge.send("batch_friend_req", {
                "userIds": actual_ids, "message": msg, "sourceGroupId": gid,
                "delayMs": delay, "imei": acc['imei'], "cookie": acc['cookie']
            }, lambda d,e: self.root.after(0, self._batch_progress_handler, d, e, acc['name'], "Kết bạn"))

        threading.Thread(target=run, daemon=True).start()

    def _batch_invite(self):
        actual_ids, acc = self._get_ids_for_action("invite")
        if not actual_ids: return
        gid = self.group_id_entry.get().strip()
        if not gid: return messagebox.showwarning("!", "Nhập ID hoặc Link nhóm!")
        delay = self._get_delay()
        limit = self._get_limit()

        def run():
            self._log(f"📨 Mời nhóm: {len(actual_ids)} người → {gid}", "info", acc['name'])
            self.bridge.send("invite_to_group", {
                "userIds": actual_ids, "groupId": gid, "delayMs": delay,
                "limit": limit,
                "imei": acc['imei'], "cookie": acc['cookie']
            }, lambda d,e: self.root.after(0, self._batch_progress_handler, d, e, acc['name'], "Mời nhóm"))

        threading.Thread(target=run, daemon=True).start()

    def _batch_message(self):
        actual_ids, acc = self._get_ids_for_action("message")
        if not actual_ids: return
        
        limit = self._get_limit()
        delay = self._get_delay()

        # Capture current state for this specific thread
        use_rand = self.use_random.get()
        sel_tpl_name = self.tpl_msg_sel.get()
        tpls = self.templates.copy()

        def run_thread():
            msg = "Chào bạn!"
            img = ""
            if use_rand:
                m_tpls = [t for t in tpls if t["type"] == "Nhắn tin"]
                if m_tpls:
                    chosen = random.choice(m_tpls)
                    msg = chosen["content"]
                    img = chosen.get("image", "")
            else:
                for t in tpls:
                    if t.get("name") == sel_tpl_name:
                        msg = t["content"]
                        img = t.get("image", "")
                        break
            
            self._log(f"💬 Gửi tin: {len(actual_ids)} người (Acc:{acc['name']})", "info", acc['name'])
            self._log(f"📝 Nội dung: '{msg[:50]}...' | Ảnh: '{img or 'không'}' | Random: {use_rand} | Tpl: '{sel_tpl_name}'", "info", acc['name'])
            
            # Lấy ID nhóm nguồn để gửi tin nhắn cho người cùng nhóm (kể cả người lạ)
            gid = self.scan_result.get("groupInfo", {}).get("groupId", "") if self.scan_result else ""
            
            def cb(d, e):
                self.root.after(0, self._batch_progress_handler, d, e, acc['name'], "Nhắn tin")
            
            self.bridge.send("batch_send_msg", {
                "userIds": actual_ids, "message": msg, "imagePath": img if img else None,
                "delayMs": delay, "limit": limit, "sourceGroupId": gid,
                "imei": acc['imei'], "cookie": acc['cookie']
            }, cb)

        threading.Thread(target=run_thread, daemon=True).start()

    # ---- Templates Manager ----
    def _load_templates(self):
        try:
            if TEMPLATES_FILE.exists():
                return json.loads(TEMPLATES_FILE.read_text("utf-8"))
        except: pass
        return [
            {"type": "Kết bạn", "content": "Xin chào! Mình muốn kết bạn."},
            {"type": "Nhắn tin", "content": "Chào bạn, mời bạn tham gia nhóm zalo của mình: https://zalo.me/g/..."}
        ]

    def _save_templates(self):
        TEMPLATES_FILE.write_text(json.dumps(self.templates, ensure_ascii=False, indent=2), "utf-8")

    def _build_template_tab(self, parent):
        # Header for the tab
        head = tk.Frame(parent, bg=C["card2"], pady=10)
        head.pack(fill="x")
        tk.Label(head, text="📜 THƯ VIỆN KỊCH BẢN CHỐNG SPAM", font=F["h2"], bg=C["card2"], fg=C["bright"]).pack()

        main = tk.Frame(parent, bg=C["bg"], padx=15, pady=15)
        main.pack(fill="both", expand=True)

        # Left: List of templates
        left = tk.Frame(main, bg=C["bg"])
        left.pack(side="left", fill="both", expand=True, padx=(0,10))
        
        make_label(left, "Danh sách kịch bản hiện có:").pack(anchor="w")
        
        # Style for Treeview in this tab
        s = ttk.Style()
        s.configure("Tpl.Treeview", rowheight=30)
        
        self.tpl_tree = ttk.Treeview(left, columns=("name", "type", "content"), show="headings", height=15, style="Tpl.Treeview")
        self.tpl_tree.heading("name", text="Tên kịch bản")
        self.tpl_tree.heading("type", text="Loại")
        self.tpl_tree.heading("content", text="Nội dung")
        self.tpl_tree.column("name", width=150, anchor="w")
        self.tpl_tree.column("type", width=100, anchor="center")
        self.tpl_tree.column("content", width=350, anchor="w")
        self.tpl_tree.pack(fill="both", expand=True)
        self.tpl_tree.bind("<<TreeviewSelect>>", self._on_tpl_select)

        # Right: Editor
        right = tk.Frame(main, bg=C["card"], padx=15, pady=15, highlightthickness=1, highlightbackground=C["border"], width=380)
        right.pack(side="right", fill="y")
        right.pack_propagate(False)
        
        make_label(right, "1. Loại kịch bản:").pack(anchor="w")
        self.tpl_type_var = tk.StringVar(value="Kết bạn")
        cb = ttk.Combobox(right, textvariable=self.tpl_type_var, 
                          values=["Kết bạn", "Nhắn tin", "Mời nhóm"], state="readonly", font=F["body"])
        cb.pack(fill="x", pady=(2,10))

        make_label(right, "2. Tên kịch bản (để gợi nhớ):").pack(anchor="w")
        self.tpl_name_entry = make_entry(right)
        self.tpl_name_entry.pack(fill="x", pady=(2,10))

        make_label(right, "3. Nội dung kịch bản:").pack(anchor="w")
        f_edit = tk.Frame(right, bg=C["input"], highlightthickness=1, highlightbackground=C["border"])
        f_edit.pack(fill="both", expand=True, pady=(2,10))
        self.tpl_edit = tk.Text(f_edit, font=F["body"], bg=C["input"], fg=C["text"], 
                                insertbackground=C["text"], height=8, relief="flat", bd=5, wrap="word")
        self.tpl_edit.pack(fill="both", expand=True)

        make_label(right, "4. Ảnh đính kèm cho kịch bản:").pack(anchor="w")
        self.tpl_image_path = tk.StringVar(value="")
        lbl_img = tk.Label(right, textvariable=self.tpl_image_path, font=F["sm"], bg=C["card"], fg=C["accent"], wraplength=350)
        lbl_img.pack(fill="x")
        
        f_img = tk.Frame(right, bg=C["card"])
        f_img.pack(fill="x", pady=5)
        Btn(f_img, "🖼️ Chọn ảnh", cmd=self._pick_tpl_image, color="#4b5563").pack(side="left", expand=True, fill="x", padx=(0,2))
        Btn(f_img, "🗑️ Xóa ảnh", cmd=self._clear_tpl_image, color=C["red"]).pack(side="left", expand=True, fill="x", padx=(2,0))

        Btn(right, "💾 LƯU KỊCH BẢN", cmd=self._save_tpl_item, color=C["green"]).pack(fill="x", pady=5)
        Btn(right, "🗑️ XÓA KỊCH BẢN CHỌN", cmd=self._delete_tpl_item, color=C["red"]).pack(fill="x", pady=5)
        
        tk.Label(right, text="💡 Gợi ý: Tạo ít nhất 3-5 mẫu kịch bản cho \nmỗi loại để tỷ lệ an toàn cao nhất.", 
                 font=F["sm"], bg=C["card"], fg=C["dim"], justify="left").pack(pady=10)
        
        self._refresh_tpl_table()

    def _on_tpl_select(self, e):
        sel = self.tpl_tree.selection()
        if not sel: return
        idx = self.tpl_tree.index(sel[0])
        t = self.templates[idx]
        self.tpl_type_var.set(t["type"])
        self.tpl_name_entry.delete(0, "end")
        self.tpl_name_entry.insert(0, t.get("name", ""))
        self.tpl_edit.delete("1.0", "end")
        self.tpl_edit.insert("1.0", t["content"])
        self.tpl_image_path.set(t.get("image", ""))

    def _save_tpl_item(self):
        t_type = self.tpl_type_var.get()
        t_name = self.tpl_name_entry.get().strip()
        t_content = self.tpl_edit.get("1.0", "end").strip()
        t_image = self.tpl_image_path.get()
        
        if not t_name or not t_content:
            messagebox.showwarning("Thiếu thông tin", "Vui lòng nhập đầy đủ Tên và Nội dung kịch bản.")
            return
        
        item = {"name": t_name, "type": t_type, "content": t_content, "image": t_image}
        
        # Check if updating
        sel = self.tpl_tree.selection()
        if sel:
            idx = self.tpl_tree.index(sel[0])
            self.templates[idx] = item
            self._log(f"✅ Đã cập nhật kịch bản '{t_name}'", "ok")
        else:
            self.templates.append(item)
            self._log(f"✅ Đã thêm kịch bản '{t_name}'", "ok")
            
        self._save_templates()
        self._refresh_tpl_table()
        self._refresh_action_combos() # Đồng bộ sang tab Vận hành
        self._clear_tpl_editor() # Clear form sau khi lưu
        self.tpl_name_entry.delete(0, "end")
        self.tpl_edit.delete("1.0", "end")
        self.tpl_image_path.set("")

    def _pick_tpl_image(self):
        fp = filedialog.askopenfilename(filetypes=[("Image Files", "*.png;*.jpg;*.jpeg;*.gif;*.bmp")])
        if fp:
            try:
                original_path = Path(fp)
                new_fn = f"tpl_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{original_path.name}"
                dest_path = IMAGE_DIR / new_fn
                shutil.copy2(fp, dest_path)
                self.tpl_image_path.set(str(dest_path.absolute()))
                self._log(f"🖼️ Đã gán ảnh cho kịch bản: {new_fn}", "ok")
            except Exception as e:
                self._log(f"❌ Lỗi gán ảnh: {e}", "error")

    def _clear_tpl_image(self):
        self.tpl_image_path.set("")

    def _delete_tpl_item(self):
        sel = self.tpl_tree.selection()
        if not sel: return
        idx = self.tpl_tree.index(sel[0])
        if messagebox.askyesno("Xác nhận", "Xóa kịch bản này?"):
            self.templates.pop(idx)
            self._save_templates()
            self._refresh_tpl_table()
            self._refresh_action_combos()
            self.tpl_edit.delete("1.0", "end")
            self._log(f"🗑️ Đã xóa kịch bản khỏi thư viện", "warn")

    def _refresh_action_combos(self):
        """Cập nhật dropdown ở Tab Vận hành - Đồng bộ ngay khi Thêm/Sửa/Xóa"""
        f_tpls = [t.get("name", "Không tên") for t in self.templates if t["type"] == "Kết bạn"]
        i_tpls = [t.get("name", "Không tên") for t in self.templates if t["type"] == "Mời nhóm"]
        m_tpls = [t.get("name", "Không tên") for t in self.templates if t["type"] == "Nhắn tin"]

        # Cập nhật Tab Vận hành: Kết bạn
        self.tpl_friend_combo["values"] = f_tpls
        curr_f = self.tpl_friend_sel.get()
        if not f_tpls:
            self.tpl_friend_sel.set("")
        elif not curr_f or curr_f not in f_tpls:
            self.tpl_friend_sel.set(f_tpls[0])

        # Cập nhật Tab Vận hành: Mời nhóm
        self.tpl_invite_combo["values"] = i_tpls
        curr_i = self.tpl_invite_sel.get()
        if not i_tpls:
            self.tpl_invite_sel.set("")
        elif not curr_i or curr_i not in i_tpls:
            self.tpl_invite_sel.set(i_tpls[0])

        # Cập nhật Tab Vận hành: Nhắn tin
        self.tpl_msg_combo["values"] = m_tpls
        curr_m = self.tpl_msg_sel.get()
        if not m_tpls:
            self.tpl_msg_sel.set("")
        elif not curr_m or curr_m not in m_tpls:
            self.tpl_msg_sel.set(m_tpls[0])

    def _refresh_tpl_table(self):
        self.tpl_tree.delete(*self.tpl_tree.get_children())
        for t in self.templates:
            content_preview = t["content"].replace("\n", " ")[:50] + "..."
            self.tpl_tree.insert("", "end", values=(t.get("name","?"), t["type"], content_preview))

    def on_close(self):
        if self.bridge: self.bridge.stop()
        self.root.destroy()


# ============================================================
if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.protocol("WM_DELETE_WINDOW", app.on_close)
    root.mainloop()
